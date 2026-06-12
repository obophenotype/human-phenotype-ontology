#!/usr/bin/env python3
"""
Extract a single named sheet from an .xlsx workbook and emit it as TSV.

Usage:
    python3 xlsx_to_tsv.py <workbook.xlsx> <sheet_name> <output.tsv>

Empty trailing columns (from over-wide xlsx ranges) are trimmed so the
resulting TSV stays compatible with ROBOT template parsing.
"""
import csv
import sys

import openpyxl


def main():
    if len(sys.argv) != 4:
        print(__doc__, file=sys.stderr)
        sys.exit(2)

    xlsx_path, sheet_name, out_tsv = sys.argv[1], sys.argv[2], sys.argv[3]

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        sys.exit(
            f"Sheet '{sheet_name}' not found in {xlsx_path}. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    def flatten(v):
        # ROBOT's template parser handles single-line TSV most reliably.
        # Replace any in-cell CR/LF with a single space so multi-line cells
        # don't split a logical row across multiple TSV lines.
        if v is None:
            return ""
        s = str(v)
        return s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")

    max_col = 0
    for r in rows:
        for i, v in enumerate(r):
            if v not in (None, ""):
                max_col = max(max_col, i + 1)

    with open(out_tsv, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh, delimiter="\t", lineterminator="\n")
        for r in rows:
            trimmed = [flatten(v) for v in r[:max_col]]
            # Skip rows that are entirely empty after trimming.
            if not any(cell.strip() for cell in trimmed):
                continue
            # Pad short rows so every line has the same field count as the
            # header — required by tsvalid and by ROBOT's template parser.
            trimmed += [""] * (max_col - len(trimmed))
            writer.writerow(trimmed)


if __name__ == "__main__":
    main()
